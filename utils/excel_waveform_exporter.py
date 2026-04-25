"""
Excel 파형 시각화 내보내기

신호 파라미터를 Excel 셀 테두리로 파형을 시각화합니다.

출력 레이아웃:
  1행 : 구간 라벨 (Seg1, Seg2, ...)
  2행 이후 : 신호별 블록
             ├ Timing행 : delay/width/period 구분 표시
             ├ H행 (High)  : 양전압 파형 상단 테두리
             ├ M행 (Mid)   : 0V 파형 테두리
             └ L행 (Low)   : 음전압 파형 하단 테두리
             └ (여백 행)

파형 그리기 규칙:
  - 전압 > 0:  H행에 top border
  - 전압 = 0:  M행에 top border
  - 전압 < 0:  L행에 bottom border
  - Edge(전환): 해당 구간의 C1 열에 H/M/L 세 행 모두 left 굵은선(수직 연결)
  - 구간 경계 점선 없음 (깔끔한 표시를 위해 제거)

Timing 표시 규칙 (파형 윗 Timing행):
  - Delay 구간: "D: Xus" (보라색)  → delay 열 범위 중앙
  - Width 구간: "W: Xus" (파란색)  → width 열 범위 중앙
  - Period:     "T: Xus" (주황색)  → period 열 범위 우측
  - 시간 단위: 항상 us (소수점 이하 반올림 없이 전부 표시)
"""

import re
from typing import List, Dict, Tuple, Optional


# ────────────────────────────────────────────────────────────────
# 레이아웃 상수
# ────────────────────────────────────────────────────────────────

ZOOM_PERCENT     = 25       # 시트 기본 줌 배율 (%)

COL_NUM          = 1        # A열: 신호 NUM
COL_NAME         = 2        # B열: 신호 이름
COL_WAVE_START   = 3        # C열 이후: 파형 셀 영역

ROW_LABEL        = 1        # 1행: 구간 라벨 (Seg1, Seg2, ...)
ROW_WAVE_START   = 3        # 3행부터 신호 파형 시작 (2행 = 첫 신호 timing행)

CELLS_PER_SIGNAL = 3        # 신호당 행 수: H행(위), M행(가운데), L행(아래)
SIGNAL_GAP_ROWS  = 2        # 신호 간 여백 행 수 (마지막 여백행 = 다음 신호 timing행)

NAME_COL_WIDTH   = 18.0     # 신호 이름 열 너비
NUM_COL_WIDTH    = 7.0      # 신호 번호 열 너비
WAVE_COL_WIDTH   = 2.5      # 파형 셀 열 너비

LABEL_COLOR      = 'FFD0D8E8'   # 구간 라벨 배경색 (연파랑)

FONT_LABEL       = 9
FONT_VOLTAGE     = 7
FONT_TIMING      = 7
FONT_NAME        = 10


# ────────────────────────────────────────────────────────────────
# 시간 포맷: 항상 us 단위, 반올림/잘림 없이 표시
# ────────────────────────────────────────────────────────────────

def _format_us(us: float) -> str:
    """시간값을 us 단위 문자열로 반환 (ms/ns 등 다른 단위 사용 금지)"""
    if us == int(us):
        return f"{int(us)}us"
    # 부동소수점 표현에서 불필요한 trailing zero 제거
    s = f"{us:.6f}".rstrip('0').rstrip('.')
    return f"{s}us"


# ────────────────────────────────────────────────────────────────
# 구간 계산
# ────────────────────────────────────────────────────────────────

def _compute_segments(sync_data_us: float, signals: List[Dict],
                      n_frames: int = 2) -> List[Tuple[float, float, str]]:
    """
    신호 타이밍 경계점 기반 구간 분할

    동작 순서:
      1. total_us = sync_data_us × n_frames (기본 2 frame 표시)
      2. 각 신호의 타이밍 경계점(delay, delay+width, period 배수 등)을
         각 프레임의 offset을 더해 수집.
         - delay/width 경계: 신호가 Low↔High로 전환되는 지점
         - period 경계: 다음 주기가 시작되는 지점 (모든 배수 추가)
         - 프레임 경계(sync_data_us, 2×sync_data_us)도 breakpoint로 추가
      3. 수집한 breakpoints를 정렬 → 인접 쌍을 구간([start, end])으로 변환
      4. 각 구간에 "Seg1", "Seg2", ... 라벨 부여 (구간 수 제한 없음)

    Args:
        sync_data_us: 1 frame 길이(us)
        signals: 신호 파라미터 딕셔너리 리스트
        n_frames: 표시할 최소 프레임 수 (기본 2)

    Returns:
        List of (start_us, end_us, label) tuples
    """
    total_us = sync_data_us * n_frames
    breakpoints = {0.0, total_us}

    for frame in range(n_frames):
        offset = frame * sync_data_us
        # 프레임 경계점 추가 (두 프레임의 경계가 명확하게 표시됨)
        if 0 < offset < total_us:
            breakpoints.add(offset)
        for sig in signals:
            delay  = float(sig.get('delay',  0))
            width  = float(sig.get('width',  0))
            period = float(sig.get('period', 0))
            # delay >= period인 경우 정규화 (OTD에서 delay가 period보다 클 수 있음)
            eff_delay = delay % period if period > 0 else delay
            if period > 0:
                # 프레임 내 모든 주기 배수의 경계점 추가 (period < sync_data_us 경우 포함)
                k = 0
                while True:
                    base = k * period
                    if offset + base >= total_us:
                        break
                    for bp in [base + eff_delay,
                               base + eff_delay + width,
                               (k + 1) * period]:
                        real_bp = offset + bp
                        if 0 < real_bp < total_us:
                            breakpoints.add(real_bp)
                    k += 1
            else:
                for bp in [eff_delay, eff_delay + width]:
                    real_bp = offset + bp
                    if 0 < real_bp < total_us:
                        breakpoints.add(real_bp)

    sorted_bps = sorted(breakpoints)
    raw = list(zip(sorted_bps[:-1], sorted_bps[1:]))
    return [(s, e, f"Seg{i+1}") for i, (s, e) in enumerate(raw)]


def _get_level(sig: Dict, t_us: float) -> float:
    """
    시간 t_us에서 신호의 전압값(V) 반환

    DC 모드 (delay=width=period=0): 항상 V1 출력
    일반 모드: delay~delay+width 구간에서 V2, 그 외 V1
    """
    v1 = float(sig.get('v1', 0))
    v2 = float(sig.get('v2', v1))
    delay  = float(sig.get('delay',  0))
    width  = float(sig.get('width',  0))
    period = float(sig.get('period', 0))

    if delay == 0 and width == 0 and period == 0:
        return v1

    if period > 0:
        eff_delay = delay % period  # delay >= period인 경우 정규화
        phase = t_us % period
        in_pulse = (eff_delay <= phase < eff_delay + width)
    else:
        in_pulse = (delay <= t_us < delay + width)

    return v2 if in_pulse else v1


def _sig_base_row(sig_idx: int) -> int:
    """신호 인덱스를 파형 시작 행 번호(H행)로 변환"""
    return ROW_WAVE_START + sig_idx * (CELLS_PER_SIGNAL + SIGNAL_GAP_ROWS)


def _sig_timing_row(sig_idx: int) -> int:
    """신호 인덱스를 timing 라벨 행 번호로 변환 (H행 바로 위)"""
    return _sig_base_row(sig_idx) - 1


# ────────────────────────────────────────────────────────────────
# 메인 익스포터
# ────────────────────────────────────────────────────────────────

class ExcelWaveformExporter:
    """
    Excel 파형 시각화 내보내기 클래스

    openpyxl을 사용하여 신호 파형을 셀 테두리로 시각화합니다.
    """

    def __init__(self):
        self._pending_shapes: Dict[str, List[str]] = {}
        self._shape_id_counter: Dict[str, int] = {}

    def export_all_models(self, filepath: str, model_store) -> bool:
        """
        model_store의 전체 모델을 각 시트에 파형으로 내보내기
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl이 필요합니다. 'pip install openpyxl'을 실행하세요.")

        if not model_store.models:
            return False

        self._pending_shapes = {}
        self._shape_id_counter = {}

        wb = Workbook()
        first_sheet = True

        for md in model_store.models:
            sig_dicts = []
            for i, sig in enumerate(md.signals):
                if hasattr(sig, 'to_dict'):
                    d = sig.to_dict()
                    d['num'] = getattr(sig, '_num', f'S{i+1:02d}')
                else:
                    d = dict(sig)
                sig_dicts.append(d)

            # 모든 값이 0인 신호(visible=False) 제외
            visible = [s for s in sig_dicts if s.get('visible', True)]
            if not visible:
                continue

            sync_data_us = md.sync_data_us
            model_name   = md.name or md.model_num

            if first_sheet:
                ws = wb.active
                safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
                ws.title = safe_title
                first_sheet = False
            else:
                safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
                ws = wb.create_sheet(title=safe_title)

            self._draw_sheet(ws, visible, sync_data_us, model_name)

        wb.save(filepath)
        self._inject_drawings_zipfile(filepath, wb)
        return True

    def export(self, filepath: str, signals: List[Dict],
               sync_data_us: float, model_name: str = "Model") -> bool:
        """단일 모델 신호 파형을 Excel 파일로 내보내기 (하위 호환용)"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl이 필요합니다. 'pip install openpyxl'을 실행하세요.")

        if not signals:
            return False

        self._pending_shapes = {}
        self._shape_id_counter = {}

        wb = Workbook()
        ws = wb.active
        safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
        ws.title = safe_title
        ws.sheet_view.zoomScale = ZOOM_PERCENT

        self._draw_sheet(ws, signals, sync_data_us, model_name)

        wb.save(filepath)
        self._inject_drawings_zipfile(filepath, wb)
        return True

    def _draw_sheet(self, ws, signals: List[Dict],
                    sync_data_us: float, model_name: str):
        """워크시트에 파형 그리기"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.sheet_view.zoomScale = ZOOM_PERCENT

        # ── 구간 계산 (최소 2 frame, 구간 수 제한 없음) ──────────
        segments = _compute_segments(sync_data_us, signals, n_frames=2)
        n_segs   = len(segments)
        total_us = sync_data_us * 2   # 2 frame 기준 전체 시간

        TOTAL_WAVE_COLS = n_segs * 10
        seg_cols = []
        for s, e, _ in segments:
            ratio = (e - s) / total_us if total_us > 0 else 1 / n_segs
            seg_cols.append(max(1, round(ratio * TOTAL_WAVE_COLS)))

        diff = TOTAL_WAVE_COLS - sum(seg_cols)
        seg_cols[-1] = max(1, seg_cols[-1] + diff)

        seg_start_cols = []
        cur = COL_WAVE_START
        for n in seg_cols:
            seg_start_cols.append(cur)
            cur += n
        total_wave_end_col = cur - 1

        # ── 열 너비 설정 ───────────────────────────────────────────
        ws.column_dimensions[get_column_letter(COL_NUM)].width  = NUM_COL_WIDTH
        ws.column_dimensions[get_column_letter(COL_NAME)].width = NAME_COL_WIDTH
        for c in range(COL_WAVE_START, total_wave_end_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = WAVE_COL_WIDTH

        # ── 행 높이 설정 ───────────────────────────────────────────
        ws.row_dimensions[ROW_LABEL].height = 22
        # 첫 신호의 timing행 (row 2) 높이
        ws.row_dimensions[ROW_WAVE_START - 1].height = 14

        for sig_idx in range(len(signals)):
            base = _sig_base_row(sig_idx)
            timing_row = _sig_timing_row(sig_idx)
            ws.row_dimensions[timing_row].height = 14
            for r in range(base, base + CELLS_PER_SIGNAL):
                ws.row_dimensions[r].height = 10
            # 여백 행 (gap의 첫 번째 행만 여백, 두 번째는 다음 신호 timing행)
            gap_row1 = base + CELLS_PER_SIGNAL
            ws.row_dimensions[gap_row1].height = 4

        # ── 스타일 헬퍼 ────────────────────────────────────────────
        center_align = Alignment(horizontal='center', vertical='center')

        def solid_side(style='thin', color='FF000000'):
            return Side(style=style, color=color)

        header_fill = PatternFill('solid', fgColor=LABEL_COLOR)
        label_font  = Font(bold=True, size=FONT_LABEL)

        # ── 행 1: 구간 라벨 ─────────────────────────────────────
        ws.cell(ROW_LABEL, COL_NUM,  "NUM").font  = label_font
        ws.cell(ROW_LABEL, COL_NAME, "신호 이름").font = label_font
        ws.cell(ROW_LABEL, COL_NUM).fill  = header_fill
        ws.cell(ROW_LABEL, COL_NAME).fill = header_fill
        ws.cell(ROW_LABEL, COL_NUM).alignment  = center_align
        ws.cell(ROW_LABEL, COL_NAME).alignment = center_align

        for si, (start, end, label) in enumerate(segments):
            c1 = seg_start_cols[si]
            c2 = c1 + seg_cols[si] - 1
            cell = ws.cell(ROW_LABEL, c1, label)
            cell.font      = label_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = Border(
                left=solid_side('medium'), right=solid_side('medium'),
                top=solid_side('medium'),  bottom=solid_side('medium'))
            if c2 > c1:
                ws.merge_cells(start_row=ROW_LABEL, start_column=c1,
                               end_row=ROW_LABEL, end_column=c2)

        # ── 신호별 파형 그리기 ────────────────────────────────────
        for sig_idx, sig in enumerate(signals):
            base  = _sig_base_row(sig_idx)
            h_row = base           # High 행
            m_row = base + 1       # Mid 행
            l_row = base + 2       # Low 행

            # NUM / NAME 셀
            num_str = sig.get('num', f'S{sig_idx + 1:02d}')
            ws.cell(h_row, COL_NUM, num_str).font = Font(bold=True, size=9)
            ws.cell(h_row, COL_NUM).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=COL_NUM,
                           end_row=l_row,   end_column=COL_NUM)

            ws.cell(h_row, COL_NAME, sig.get('name', '')).font = Font(bold=True, size=FONT_NAME)
            ws.cell(h_row, COL_NAME).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=COL_NAME,
                           end_row=l_row,   end_column=COL_NAME)

            # 구간별 파형 셀 그리기
            prev_voltage = None
            for si_idx, (start, end, _) in enumerate(segments):
                mid_t   = (start + end) / 2
                voltage = _get_level(sig, mid_t)
                c1 = seg_start_cols[si_idx]
                c2 = c1 + seg_cols[si_idx] - 1

                is_transition = (prev_voltage is not None and
                                 abs(prev_voltage - voltage) > 1e-6)

                self._draw_waveform_cells(
                    ws, h_row, m_row, l_row, c1, c2,
                    voltage, prev_voltage, is_transition,
                    solid_side
                )

                # 전압값 레이블: 첫 구간 또는 파형 전환 구간에 표시
                if si_idx == 0 or is_transition:
                    v_label = f"{voltage:+.1f}V"
                    if voltage == 0:
                        vc = ws.cell(m_row, c1, v_label)
                        vc.font = Font(color='FF00AA00', size=FONT_VOLTAGE, bold=True)
                    elif voltage > 0:
                        vc = ws.cell(h_row, c1, v_label)
                        vc.font = Font(color='FFCC0000', size=FONT_VOLTAGE, bold=True)
                    else:
                        vc = ws.cell(l_row, c1, v_label)
                        vc.font = Font(color='FF0033CC', size=FONT_VOLTAGE, bold=True)

                prev_voltage = voltage

            # timing 화살표 선 도형 수집 (openpyxl drawing으로 추가)
            self._collect_timing_shapes(
                ws.title, sig, sig_idx, segments, seg_start_cols, seg_cols, total_us
            )

    def _draw_waveform_cells(self, ws, h_row, m_row, l_row,
                              c1, c2, voltage, prev_voltage,
                              is_transition, solid_side_fn):
        """
        파형 셀 테두리 그리기

        - 전환(transition) 구간의 c1 열: 이전/현재 전압 레벨 행 사이만 굵은 left border
          RT(상승): 이전 행→현재 행 (아래→위) 구간만 / FT(하강): 반대
        - 비전환 구간 경계에는 border 없음 (깔끔한 파형 표시)
        - 수평 파형선: 전압 레벨에 따라 H/M/L 중 해당 행에만 그림
        """
        from openpyxl.styles import Border, Side

        thick = solid_side_fn('thick')
        none  = Side(style=None)

        def _vol_row(v):
            if v > 0: return h_row
            elif v == 0: return m_row
            else: return l_row

        for col in range(c1, c2 + 1):
            is_left_col = (col == c1)

            # 좌측 경계선: 출발 행 + 중간 행(zero crossing 시), 목적지 행 제외
            # 규칙: L→M: L행, L→H: L+M행, H→M: H행, H→L: H+M행, M↔H/L: M행
            if is_transition and is_left_col and prev_voltage is not None:
                prev_r = _vol_row(prev_voltage)
                curr_r = _vol_row(voltage)
                border_rows = {prev_r}
                # 전압이 양과 음 사이를 직접 교차하면 중간(M) 행도 추가
                if (prev_r == h_row and curr_r == l_row) or \
                   (prev_r == l_row and curr_r == h_row):
                    border_rows.add(m_row)
                left_h = thick if h_row in border_rows else none
                left_m = thick if m_row in border_rows else none
                left_l = thick if l_row in border_rows else none
            else:
                left_h = none
                left_m = none
                left_l = none

            # 수평 파형선
            if voltage > 0:
                ws.cell(h_row, col).border = Border(top=thick,    left=left_h)
                ws.cell(m_row, col).border = Border(left=left_m)
                ws.cell(l_row, col).border = Border(left=left_l)
            elif voltage == 0:
                ws.cell(h_row, col).border = Border(left=left_h)
                ws.cell(m_row, col).border = Border(top=thick,    left=left_m)
                ws.cell(l_row, col).border = Border(left=left_l)
            else:
                ws.cell(h_row, col).border = Border(left=left_h)
                ws.cell(m_row, col).border = Border(left=left_m)
                ws.cell(l_row, col).border = Border(bottom=thick, left=left_l)

    # ── 타이밍 화살표 도형 ───────────────────────────────────────────

    def _collect_timing_shapes(self, sheet_title: str, sig: Dict, sig_idx: int,
                                segments, seg_start_cols, seg_cols,
                                sync_data_us: float) -> None:
        """
        타이밍 구간별 선 화살표 + 텍스트 상자 도형 XML을 수집.
        _apply_shapes_to_worksheets 에서 openpyxl drawing으로 추가.

        구간:
          ① Start → 첫 RT : Delay (보라)
          ② RT → FT       : Width (파랑)
          ③ FT → 다음 RT  : Period - Width (주황)
        """
        delay  = float(sig.get('delay',  0))
        width  = float(sig.get('width',  0))
        period = float(sig.get('period', 0))

        if delay == 0 and width == 0 and period == 0:
            return

        eff_delay = delay % period if period > 0 else delay
        timing_row_0 = _sig_timing_row(sig_idx) - 1   # 0-based row
        total_cols   = sum(seg_cols)

        def us_to_col_0(us_val: float) -> int:
            # 세그먼트 경계 기반 열 계산 (반올림 오차 없이 정확한 경계 스냅)
            for i, (s, e, _) in enumerate(segments):
                if abs(s - us_val) < 1e-6:
                    return seg_start_cols[i] - 1  # 0-based
                if s < us_val < e:
                    frac = (us_val - s) / (e - s)
                    return seg_start_cols[i] - 1 + int(frac * seg_cols[i])
            return seg_start_cols[-1] + seg_cols[-1] - 1  # 마지막 경계 (0-based)

        intervals = []
        if eff_delay > 0:
            intervals.append((0.0, eff_delay, '7030A0', f"D: {_format_us(eff_delay)}"))
        if width > 0:
            intervals.append((eff_delay, eff_delay + width, '0070C0', f"W: {_format_us(width)}"))
        if period > 0:
            rest = period - eff_delay - width
            if rest > 0:
                intervals.append((eff_delay + width, period, 'CC5500', f"P: {_format_us(period)}"))

        if not intervals:
            return

        if sheet_title not in self._pending_shapes:
            self._pending_shapes[sheet_title]    = []
            self._shape_id_counter[sheet_title]  = 1

        # 행 높이 절반 EMU (기본 행 높이 ≈ 190500 EMU)
        HALF_ROW_EMU = 95250

        for us_start, us_end, color, label in intervals:
            c0 = us_to_col_0(us_start)
            c1 = us_to_col_0(us_end)
            if c1 <= c0:
                c1 = c0 + 1

            sid = self._shape_id_counter[sheet_title]

            # ── 타이밍 화살표: 선 커넥터(양방향 화살표) ──────────
            # xdr:cxnSp 를 사용하여 선 객체 생성, 양쪽 끝에 화살표 설정
            arrow = (
                '<xdr:twoCellAnchor editAs="oneCell">'
                f'<xdr:from><xdr:col>{c0}</xdr:col><xdr:colOff>0</xdr:colOff>'
                f'<xdr:row>{timing_row_0}</xdr:row><xdr:rowOff>{HALF_ROW_EMU}</xdr:rowOff></xdr:from>'
                f'<xdr:to><xdr:col>{c1}</xdr:col><xdr:colOff>0</xdr:colOff>'
                f'<xdr:row>{timing_row_0}</xdr:row><xdr:rowOff>{HALF_ROW_EMU + 1}</xdr:rowOff></xdr:to>'
                f'<xdr:cxnSp macro="">'
                f'<xdr:nvCxnSpPr>'
                f'<xdr:cNvPr id="{sid}" name="Line{sid}"/>'
                f'<xdr:cNvCxnSpPr/>'
                f'</xdr:nvCxnSpPr>'
                f'<xdr:spPr>'
                f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="1" cy="1"/></a:xfrm>'
                f'<a:prstGeom prst="line"><a:avLst/></a:prstGeom>'
                f'<a:ln w="25400">'
                f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
                f'<a:headEnd type="arrow"/>'
                f'<a:tailEnd type="arrow"/>'
                f'</a:ln>'
                f'</xdr:spPr>'
                f'</xdr:cxnSp>'
                f'<xdr:clientData/></xdr:twoCellAnchor>'
            )
            self._pending_shapes[sheet_title].append(arrow)
            self._shape_id_counter[sheet_title] += 1
            sid += 1

            # ── 텍스트 상자 (타이밍 행 상단 절반, 선 위에 라벨) ──
            textbox = (
                '<xdr:twoCellAnchor editAs="oneCell">'
                f'<xdr:from><xdr:col>{c0}</xdr:col><xdr:colOff>0</xdr:colOff>'
                f'<xdr:row>{timing_row_0}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
                f'<xdr:to><xdr:col>{c1}</xdr:col><xdr:colOff>0</xdr:colOff>'
                f'<xdr:row>{timing_row_0}</xdr:row><xdr:rowOff>{HALF_ROW_EMU}</xdr:rowOff></xdr:to>'
                f'<xdr:sp macro="" textlink=""><xdr:nvSpPr>'
                f'<xdr:cNvPr id="{sid}" name="TextBox{sid}"/>'
                f'<xdr:cNvSpPr txBox="1"><a:spLocks noGrp="1"/></xdr:cNvSpPr></xdr:nvSpPr>'
                f'<xdr:spPr>'
                f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="1" cy="1"/></a:xfrm>'
                f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom>'
                f'<a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>'
                f'<xdr:txBody><a:bodyPr anchor="ctr"/><a:lstStyle/>'
                f'<a:p><a:pPr algn="ctr"/>'
                f'<a:r><a:rPr sz="700" b="1" dirty="0">'
                f'<a:solidFill><a:srgbClr val="{color}"/></a:solidFill>'
                f'</a:rPr><a:t>{label}</a:t></a:r></a:p>'
                f'</xdr:txBody></xdr:sp><xdr:clientData/></xdr:twoCellAnchor>'
            )
            self._pending_shapes[sheet_title].append(textbox)
            self._shape_id_counter[sheet_title] += 1

    def _inject_drawings_zipfile(self, filepath: str, wb) -> None:
        """
        wb.save() 이후 xlsx를 zipfile로 열어 타이밍 화살표 도형을 주입.

        SpreadsheetDrawing.__bool__ 이 charts/images만 검사하여 빈 drawing을
        save 시 무시하므로, save 이후 직접 drawing XML, rels, Content_Types를 추가.
        """
        import zipfile
        import os

        if not self._pending_shapes:
            return

        XDR_NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
        A_NS   = 'http://schemas.openxmlformats.org/drawingml/2006/main'
        R_NS_OFF = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
        PKG_NS  = 'http://schemas.openxmlformats.org/package/2006/relationships'
        DR_TYPE = f'{R_NS_OFF}/drawing'
        DR_CT   = 'application/vnd.openxmlformats-officedocument.drawing+xml'

        # 주입할 시트 목록: (ws.path, title, drawing_index)
        to_inject = []
        drawing_idx = 0
        for ws in wb.worksheets:
            if ws.title in self._pending_shapes:
                drawing_idx += 1
                to_inject.append((ws.path, ws.title, drawing_idx))

        if not to_inject:
            return

        tmp = filepath + '._tmp'
        try:
            with zipfile.ZipFile(filepath, 'r') as zin:
                all_names = set(zin.namelist())
                files = {name: zin.read(name) for name in all_names}

            # 각 시트에 drawing 주입
            ct_overrides = []
            for ws_path, title, didx in to_inject:
                shapes = self._pending_shapes[title]
                sheet_file   = ws_path.lstrip('/')          # 'xl/worksheets/sheet1.xml'
                sheet_name   = os.path.basename(sheet_file) # 'sheet1.xml'
                rels_file    = f'xl/worksheets/_rels/{sheet_name}.rels'
                drawing_file = f'xl/drawings/drawing{didx}.xml'
                rel_id       = f'rId_d{didx}'

                # ① drawing XML
                drawing_xml = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    f'<xdr:wsDr xmlns:xdr="{XDR_NS}" xmlns:a="{A_NS}" xmlns:r="{R_NS_OFF}">'
                    + ''.join(shapes)
                    + '</xdr:wsDr>'
                ).encode('utf-8')
                files[drawing_file] = drawing_xml

                # ② worksheet rels (기존 파일이 있으면 Relationship 추가, 없으면 신규 생성)
                new_rel = (f'<Relationship Id="{rel_id}" '
                           f'Type="{DR_TYPE}" '
                           f'Target="../drawings/drawing{didx}.xml"/>')
                if rels_file in files:
                    rels_str = files[rels_file].decode('utf-8')
                    rels_str = rels_str.replace('</Relationships>', new_rel + '</Relationships>')
                else:
                    rels_str = (
                        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        f'<Relationships xmlns="{PKG_NS}">{new_rel}</Relationships>'
                    )
                files[rels_file] = rels_str.encode('utf-8')

                # ③ 시트 XML에 <drawing r:id="..."/> 삽입 (</worksheet> 직전)
                sheet_str = files[sheet_file].decode('utf-8')
                # openpyxl이 xmlns:r를 생략하는 경우 직접 선언 추가
                if 'xmlns:r=' not in sheet_str[:600]:
                    sheet_str = sheet_str.replace(
                        '<worksheet ',
                        f'<worksheet xmlns:r="{R_NS_OFF}" '
                    )
                drawing_tag = f'<drawing r:id="{rel_id}"/>'
                if drawing_tag not in sheet_str:
                    sheet_str = sheet_str.replace('</worksheet>',
                                                  drawing_tag + '</worksheet>')
                files[sheet_file] = sheet_str.encode('utf-8')

                # ④ Content_Types.xml용 Override 수집
                ct_overrides.append(
                    f'<Override PartName="/{drawing_file}" ContentType="{DR_CT}"/>'
                )

            # ④ Content_Types.xml 갱신
            ct_str = files['[Content_Types].xml'].decode('utf-8')
            for override in ct_overrides:
                if override not in ct_str:
                    ct_str = ct_str.replace('</Types>', override + '</Types>')
            files['[Content_Types].xml'] = ct_str.encode('utf-8')

            # 전체 파일을 새 zipfile로 저장
            with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
                for name, data in files.items():
                    zout.writestr(name, data)

            os.replace(tmp, filepath)

        except Exception:
            if os.path.exists(tmp):
                os.unlink(tmp)
            raise
