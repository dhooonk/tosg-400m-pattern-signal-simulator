"""
Excel PG Signal 임포터 + 양식(Format) 파일 생성기
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
포함 함수:
  - import_excel_all_models()    : Excel → ModelData 리스트 변환
  - generate_excel_format_file() : 빈 입력 양식 .xlsx 파일 생성

Excel 시트 양식 규격
─────────────────────
[신호 데이터 영역] 1~37행
  A1=NUM,  B1=NAME,  C1=SIG TYPE,  D1=SIG MODE,  E1=INV
  F1=V1(V), G1=V2(V), H1=V3(V), I1=V4(V)
  J1=DELAY(us), K1=PERIOD(us), L1=WIDTH(us), M1=LENGTH(us), N1=AREA(us)
  A2~A37 : S01~S36 (최대 36 신호)

[SyncData 영역] P/Q열, 2~4행
  P2=SyncData(us),  Q2=값
  P3=Frequency(Hz), Q3=값
  P4=SyncCounter,   Q4=값

[패턴 데이터 영역] 39행 이후
  A39= '=== PATTERN DATA ===' 구분자
  A40=PTN_NO, B40=NAME
  C~F=R_V1..V4, G~J=G_V1..V4, K~N=B_V1..V4, O~R=W_V1..V4, S40=TYPE
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[통합] 구 excel_format_generator.py → 이 파일 하단에 병합됨
"""

import os
from typing import List


# ════════════════════════════════════════════════════════
# Excel 불러오기
# ════════════════════════════════════════════════════════

def import_excel_all_models(filepath: str) -> list:
    """
    Excel 파일 전체 시트를 읽어 ModelData 리스트 반환.

    각 시트가 하나의 모델로 변환됩니다.
    시트 이름 → 모델 이름, 시트 순서(1-based 3자리) → 모델 번호.

    처리 절차:
      1. Q2(SyncData), Q3(Freq), Q4(SyncCounter) 읽기
      2. A2~A37 행에서 'S'로 시작하는 신호 행 파싱
      3. '=== PATTERN DATA ===' 구분자 이후 패턴 데이터 파싱
      4. ModelData 생성

    Args:
        filepath: .xlsx 파일 경로

    Returns:
        List[ModelData]: 파싱된 모델 데이터 리스트
                         (ModelStore.set_models()에 직접 사용 가능)

    Raises:
        ImportError       : openpyxl이 설치되지 않은 경우
        FileNotFoundError : 파일이 없는 경우
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl이 필요합니다. 'pip install openpyxl'을 실행하세요.")

    from model_store import ModelData
    from signal_model import Signal

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result_models: List[ModelData] = []

    # 신호 색상 기본 팔레트 (matplotlib 기본 10색 + 보조색)
    default_colors = [
        '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
        '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
        '#aec7e8', '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5',
    ]

    for sheet_idx, ws in enumerate(wb.worksheets):
        model_name = ws.title              # 시트 이름 = 모델 이름
        model_num  = f"{sheet_idx + 1:03d}"  # 001, 002, ...

        # ── 안전한 셀 값 변환 헬퍼 ────────────────────────────
        def _sf(v, df=0.0):
            """셀 값 → float (실패 시 df 반환)"""
            try:
                return float(v) if v is not None else df
            except (ValueError, TypeError):
                return df

        def _si(v, df=0):
            """셀 값 → int (실패 시 df 반환)"""
            try:
                return int(v) if v is not None else df
            except (ValueError, TypeError):
                return df

        # ── SyncData / Frequency / SyncCounter 읽기 ───────────
        # P열: 라벨, Q열: 값
        sync_data_us = _sf(ws['Q2'].value)   # SyncData (us)
        freq_hz      = _sf(ws['Q3'].value)   # Frequency (Hz)
        sync_cntr    = _si(ws['Q4'].value)   # SyncCounter

        # 한쪽 값이 없으면 다른 값으로 상호 계산
        if freq_hz <= 0 and sync_data_us > 0:
            freq_hz = 1_000_000.0 / sync_data_us
        elif sync_data_us <= 0 and freq_hz > 0:
            sync_data_us = 1_000_000.0 / freq_hz
        # 둘 다 없으면 기본값 60Hz 사용
        if freq_hz <= 0:
            freq_hz = 60.0
        if sync_data_us <= 0:
            sync_data_us = 1_000_000.0 / freq_hz

        # ── 신호 데이터 읽기 (2~37행, S01~S36) ───────────────
        signals: List[Signal] = []

        for row_idx in range(2, 38):
            row = ws[row_idx]
            # A열(NUM)이 비어 있으면 건너뜀
            if not row or row[0].value is None:
                continue

            num_val = str(row[0].value).strip()
            # A열이 'S'로 시작하는 행만 신호 행으로 처리 (헤더·기타 행 제외)
            if not num_val.upper().startswith('S'):
                continue

            # B열: NAME, 비어 있으면 NUM(S01 등)을 이름으로 대체
            name_val = (str(row[1].value).strip()
                        if len(row) > 1 and row[1].value else '')
            if not name_val:
                name_val = num_val

            # 열 인덱스로 float/int 읽기 헬퍼
            def col(i, dv=0.0):
                """i번째 열(0-based) 값을 float로"""
                v = row[i].value if len(row) > i else None
                try:
                    return float(v) if v is not None else dv
                except (ValueError, TypeError):
                    return dv

            def icol(i, dv=0):
                """i번째 열(0-based) 값을 int로"""
                v = row[i].value if len(row) > i else None
                try:
                    return int(v) if v is not None else dv
                except (ValueError, TypeError):
                    return dv

            # 신호 인덱스 → 색상 팔레트 순환 할당 (S01→0, S02→1, ...)
            try:
                num_idx = int(num_val[1:]) - 1
            except (ValueError, IndexError):
                num_idx = row_idx - 2
            color = default_colors[num_idx % len(default_colors)]

            # 열 순서 (0-based):
            #  0=NUM, 1=NAME, 2=SIGTYPE, 3=SIGMODE, 4=INV
            #  5=V1,  6=V2,   7=V3,      8=V4
            #  9=DELAY, 10=PERIOD, 11=WIDTH, 12=LENGTH, 13=AREA
            sig = Signal(
                name      = name_val,
                sig_type  = str(icol(2, 0)),  # C: SIG TYPE
                sig_mode  = icol(3, 0),        # D: SIG MODE
                inversion = icol(4, 0),        # E: INV
                v1 = col(5),                   # F: V1 (V)
                v2 = col(6),                   # G: V2 (V)
                v3 = col(7),                   # H: V3 (V)
                v4 = col(8),                   # I: V4 (V)
                delay  = col(9),               # J: DELAY (us)
                period = col(10),              # K: PERIOD (us)
                width  = col(11),              # L: WIDTH (us)
                color   = color,
                visible = True,
            )
            # 확장 필드: OTD 내보내기 시 LENGTH/AREA 복원에 사용
            sig._num    = num_val
            sig._length = col(12)              # M: LENGTH (us)
            sig._area   = col(13)              # N: AREA (us)
            signals.append(sig)

        # ── 패턴 데이터 읽기 ('=== PATTERN DATA ===' 구분자 이후) ──
        patterns = []
        ptn_start_row = None

        # 39행부터 구분자 행 탐색 (구분자 행+2 = 첫 데이터 행)
        for row_idx in range(39, ws.max_row + 1):
            cell_a = ws.cell(row_idx, 1).value
            if cell_a is not None and str(cell_a).strip().startswith('==='):
                ptn_start_row = row_idx + 2   # 구분자+1=헤더, +2=데이터
                break

        if ptn_start_row:
            for row_idx in range(ptn_start_row, ws.max_row + 1):
                # A~S열(1~19번) 값 읽기
                row_vals = [ws.cell(row_idx, c).value for c in range(1, 20)]
                if row_vals[0] is None:
                    continue
                try:
                    ptn_no = int(row_vals[0])
                except (ValueError, TypeError):
                    continue

                ptn_name = (str(row_vals[1]).strip()
                            if row_vals[1] else f'PTN{ptn_no:02d}')

                def pv(i):
                    """패턴 행의 i번째 값(0-based)을 float으로"""
                    try:
                        return float(row_vals[i]) if row_vals[i] is not None else 0.0
                    except (ValueError, TypeError):
                        return 0.0

                def ptype():
                    """S열(18번째, 0-based) 타입 값을 int로"""
                    try:
                        return int(row_vals[18]) if row_vals[18] is not None else 0
                    except (ValueError, TypeError):
                        return 0

                # C~F = R_V1..V4, G~J = G_V1..V4, K~N = B_V1..V4,
                # O~R = W_V1..V4, S = TYPE  (all in V)
                patterns.append({
                    'ptn_no': ptn_no,
                    'name':   ptn_name,
                    'r_v1': pv(2),  'r_v2': pv(3),  'r_v3': pv(4),  'r_v4': pv(5),
                    'g_v1': pv(6),  'g_v2': pv(7),  'g_v3': pv(8),  'g_v4': pv(9),
                    'b_v1': pv(10), 'b_v2': pv(11), 'b_v3': pv(12), 'b_v4': pv(13),
                    'w_v1': pv(14), 'w_v2': pv(15), 'w_v3': pv(16), 'w_v4': pv(17),
                    'ptn_type': ptype(),
                })

        # ── ModelData 생성 및 결과 리스트에 추가 ────────────────
        from model_store import ModelData
        result_models.append(ModelData(
            model_num    = model_num,
            name         = model_name,
            frequency_hz = freq_hz,
            sync_data_us = sync_data_us,
            sync_cntr    = sync_cntr,
            signals      = signals,
            patterns     = patterns,
        ))

    wb.close()
    return result_models


# ════════════════════════════════════════════════════════
# Excel 양식 파일 생성기
# [통합] 구 excel_format_generator.py에서 이동
# ════════════════════════════════════════════════════════

def generate_excel_format_file(filepath: str, model_count: int = 1) -> bool:
    """
    Excel 불러오기용 빈 양식(.xlsx) 파일 생성.

    import_excel_all_models()가 읽을 수 있는 올바른 구조의
    빈 Excel 파일을 생성합니다.
    사용자가 값을 채운 뒤 불러오기하면 됩니다.

    시트 구조:
      • 1행   : 신호 파라미터 헤더 (NUM, NAME, V1~V4, DELAY 등)
      • 2~37행 : 신호 데이터 입력 행 S01~S36
      • P/Q열 : SyncData(us), Frequency(Hz), SyncCounter 입력
      • 39행  : '=== PATTERN DATA ===' 구분자
      • 40행  : 패턴 헤더 (PTN_NO, NAME, R_V1~W_V4, TYPE)
      • 41~60행: 패턴 데이터 입력 행

    Args:
        filepath   : 저장할 .xlsx 파일 경로
        model_count: 생성할 시트(모델) 수 (기본: 1)

    Returns:
        bool: 저장 성공 여부
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl이 필요합니다. 'pip install openpyxl'을 실행하세요.")

    wb = Workbook()
    wb.remove(wb.active)   # 기본 시트 제거

    # ── 공통 스타일 정의 ──────────────────────────────────
    header_fill  = PatternFill('solid', fgColor='FF2C3E50')   # 진한 남색
    sync_fill    = PatternFill('solid', fgColor='FF27AE60')   # 초록
    pattern_fill = PatternFill('solid', fgColor='FF8E44AD')   # 보라
    row_fill     = PatternFill('solid', fgColor='FFECF0F1')   # 연회색 (짝수 행)

    header_font  = Font(bold=True, size=10, color='FFFFFFFF')
    sync_font    = Font(bold=True, size=9,  color='FFFFFFFF')
    pattern_font = Font(bold=True, size=9,  color='FFFFFFFF')
    num_font     = Font(bold=True, size=9,  color='FF2C3E50')

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    def _side(style='thin'):
        return Side(style=style)

    thin_border = Border(
        left=_side(), right=_side(), top=_side(), bottom=_side()
    )

    # ── 신호 헤더 컬럼 정의 (열문자, 라벨, 너비) ─────────────
    SIGNAL_HEADERS = [
        ('A', 'NUM',       8),  ('B', 'NAME',      14),
        ('C', 'SIG TYPE',  9),  ('D', 'SIG MODE',   9),
        ('E', 'INV',       7),
        ('F', 'V1 (V)',    9),  ('G', 'V2 (V)',     9),
        ('H', 'V3 (V)',    9),  ('I', 'V4 (V)',     9),
        ('J', 'DELAY (us)',  11), ('K', 'PERIOD (us)', 11),
        ('L', 'WIDTH (us)',  11), ('M', 'LENGTH (us)', 11),
        ('N', 'AREA (us)',   11),
    ]

    # ── 패턴 헤더 컬럼 정의 ─────────────────────────────────
    PATTERN_HEADERS = [
        ('A', 'PTN_NO', 8), ('B', 'NAME',  12),
        ('C', 'R_V1',   8), ('D', 'R_V2',   8),
        ('E', 'R_V3',   8), ('F', 'R_V4',   8),
        ('G', 'G_V1',   8), ('H', 'G_V2',   8),
        ('I', 'G_V3',   8), ('J', 'G_V4',   8),
        ('K', 'B_V1',   8), ('L', 'B_V2',   8),
        ('M', 'B_V3',   8), ('N', 'B_V4',   8),
        ('O', 'W_V1',   8), ('P', 'W_V2',   8),
        ('Q', 'W_V3',   8), ('R', 'W_V4',   8),
        ('S', 'TYPE',   8),
    ]

    # ── 시트별 양식 생성 ──────────────────────────────────
    for sheet_idx in range(1, model_count + 1):
        ws = wb.create_sheet(title=f"Model-{sheet_idx:03d}")

        # 행 높이 설정
        ws.row_dimensions[1].height  = 20   # 헤더 행
        ws.row_dimensions[38].height = 6    # 신호/패턴 구분 여백
        ws.row_dimensions[39].height = 18   # 패턴 구분자 행
        ws.row_dimensions[40].height = 18   # 패턴 헤더 행

        # ── 1행: 신호 파라미터 헤더 ──────────────────────────
        for col_letter, label, width in SIGNAL_HEADERS:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = thin_border
            ws.column_dimensions[col_letter].width = width

        # ── 2~37행: 신호 데이터 입력 행 (S01~S36) ────────────
        for i in range(36):
            row = i + 2
            # A열: NUM 값 (S01, S02, ...)
            num_cell = ws.cell(row=row, column=1, value=f'S{i+1:02d}')
            num_cell.font      = num_font
            num_cell.alignment = center
            num_cell.border    = thin_border
            if row % 2 == 0:
                num_cell.fill = row_fill   # 짝수 행 배경색 적용

            # B~N열: 빈 입력 셀
            for col_idx in range(2, 15):
                c = ws.cell(row=row, column=col_idx, value='')
                c.border    = thin_border
                c.alignment = center
                if row % 2 == 0:
                    c.fill = row_fill

        # ── P/Q열: SyncData 입력 영역 ─────────────────────────
        sync_labels = ['SyncData (us)', 'Frequency (Hz)', 'SyncCounter']
        for i, label in enumerate(sync_labels):
            row = i + 2
            # P열: 라벨
            p_cell = ws.cell(row=row, column=16, value=label)
            p_cell.font      = sync_font
            p_cell.fill      = sync_fill
            p_cell.alignment = left
            p_cell.border    = thin_border
            ws.column_dimensions['P'].width = 16
            # Q열: 값 입력 셀 (기본값 0)
            q_cell = ws.cell(row=row, column=17, value=0)
            q_cell.border    = thin_border
            q_cell.alignment = center
            ws.column_dimensions['Q'].width = 12

        # ── 39행: 패턴 섹션 구분자 ──────────────────────────────
        ws.merge_cells('A39:S39')
        sep = ws.cell(row=39, column=1,
                      value='=== PATTERN DATA (아래에 패턴 입력) ===')
        sep.font      = pattern_font
        sep.fill      = pattern_fill
        sep.alignment = center

        # ── 40행: 패턴 헤더 ─────────────────────────────────────
        for col_letter, label, width in PATTERN_HEADERS:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=40, column=col_idx, value=label)
            cell.font      = pattern_font
            cell.fill      = pattern_fill
            cell.alignment = center
            cell.border    = thin_border

        # ── 41~60행: 패턴 데이터 입력 행 ───────────────────────
        for ptn_row in range(41, 61):
            for col_idx in range(1, 20):
                c = ws.cell(row=ptn_row, column=col_idx, value='')
                c.border    = thin_border
                c.alignment = center
                if ptn_row % 2 == 0:
                    c.fill = row_fill
            # A열: 패턴 번호 기본값 자동 채우기
            ptn_num = ptn_row - 40
            ws.cell(row=ptn_row, column=1, value=ptn_num).font = num_font

    wb.save(filepath)
    return True
