"""
Excel 파형 시각화 내보내기

신호 파라미터를 Excel 셀 테두리로 파형을 시각화합니다.

출력 레이아웃:
  1행 : 구간 라벨 (Seg1, Seg2, ...)
  2행 이후 : 신호별 블록
             ├ Timing행 : delay/width/period 구분 표시
             ├ H행 (High)  : 양전압 파형 상단 테두리
             ├ M행 (Mid)   : 0V 파형 테두리
             └ L행 (Low)   : 음전압 파형 하단 테두리
             └ (여백 행)

파형 그리기 규칙:
  - 전압 > 0:  H행에 top border
  - 전압 = 0:  M행에 top border
  - 전압 < 0:  L행에 bottom border
  - Edge(전환): 해당 구간의 C1 열에 H/M/L 세 행 모두 left 굵은선(수직 연결)
  - 구간 경계 점선 없음 (깔끔한 표시를 위해 제거)

Timing 표시 규칙 (파형 윗 Timing행):
  - Delay 구간: "D: Xus" (보라색)  → delay 열 범위 중앙
  - Width 구간: "W: Xus" (파란색)  → width 열 범위 중앙
  - Period:     "T: Xus" (주황색)  → period 열 범위 우측
  - 시간 단위: 항상 us (소수점 이하 반올림 없이 전부 표시)
"""

import re
from typing import List, Dict, Tuple, Optional


# ────────────────────────────────────────────────────────────────
# 레이아웃 상수
# ────────────────────────────────────────────────────────────────

ZOOM_PERCENT     = 25       # 시트 기본 줌 배율 (%)

COL_NUM          = 1        # A열: 신호 NUM
COL_NAME         = 2        # B열: 신호 이름
COL_WAVE_START   = 3        # C열 이후: 파형 셀 영역

ROW_LABEL        = 1        # 1행: 구간 라벨 (Seg1, Seg2, ...)
ROW_WAVE_START   = 3        # 3행부터 신호 파형 시작 (2행 = 첫 신호 timing행)

CELLS_PER_SIGNAL = 3        # 신호당 행 수: H행(위), M행(가운데), L행(아래)
SIGNAL_GAP_ROWS  = 2        # 신호 간 여백 행 수 (마지막 여백행 = 다음 신호 timing행)

NAME_COL_WIDTH   = 18.0     # 신호 이름 열 너비
NUM_COL_WIDTH    = 7.0      # 신호 번호 열 너비
WAVE_COL_WIDTH   = 2.5      # 파형 셀 열 너비

LABEL_COLOR      = 'FFD0D8E8'   # 구간 라벨 배경색 (연파랑)

FONT_LABEL       = 9
FONT_VOLTAGE     = 7
FONT_TIMING      = 7
FONT_NAME        = 10


# ────────────────────────────────────────────────────────────────
# 시간 포맷: 항상 us 단위, 반올림/잘림 없이 표시
# ────────────────────────────────────────────────────────────────

def _format_us(us: float) -> str:
    """시간값을 us 단위 문자열로 반환 (ms/ns 등 다른 단위 사용 금지)"""
    if us == int(us):
        return f"{int(us)}us"
    # 부동소수점 표현에서 불필요한 trailing zero 제거
    s = f"{us:.6f}".rstrip('0').rstrip('.')
    return f"{s}us"


# ────────────────────────────────────────────────────────────────
# 구간 계산
# ────────────────────────────────────────────────────────────────

def _compute_segments(sync_data_us: float, signals: List[Dict],
                      max_segs: int = 32) -> List[Tuple[float, float, str]]:
    """
    신호 타이밍 경계점 기반 구간 분할

    모든 신호의 delay, width, period 경계점을 수집하여 구간을 나눕니다.
    구간이 max_segs를 초과하면 가장 짧은 인접 구간을 병합합니다.
    """
    breakpoints = {0.0, sync_data_us}
    for sig in signals:
        delay  = float(sig.get('delay',  0))
        width  = float(sig.get('width',  0))
        period = float(sig.get('period', 0))
        for bp in [delay, delay + width, period, period + delay,
                   period + delay + width]:
            if 0 < bp < sync_data_us:
                breakpoints.add(bp)

    sorted_bps = sorted(breakpoints)
    raw = list(zip(sorted_bps[:-1], sorted_bps[1:]))

    while len(raw) > max_segs:
        durs = [e - s for s, e in raw]
        idx  = durs.index(min(durs))
        if idx + 1 < len(raw):
            s1, _ = raw[idx]
            _, e2 = raw[idx + 1]
            raw = raw[:idx] + [(s1, e2)] + raw[idx + 2:]
        else:
            break

    return [(s, e, f"Seg{i+1}") for i, (s, e) in enumerate(raw)]


def _get_level(sig: Dict, t_us: float) -> float:
    """
    시간 t_us에서 신호의 전압값(V) 반환

    DC 모드 (delay=width=period=0): 항상 V1 출력
    일반 모드: delay~delay+width 구간에서 V2, 그 외 V1
    """
    v1 = float(sig.get('v1', 0))
    v2 = float(sig.get('v2', v1))
    delay  = float(sig.get('delay',  0))
    width  = float(sig.get('width',  0))
    period = float(sig.get('period', 0))

    if delay == 0 and width == 0 and period == 0:
        return v1

    if period > 0:
        phase = t_us % period
        in_pulse = (delay <= phase < delay + width)
    else:
        in_pulse = (delay <= t_us < delay + width)

    return v2 if in_pulse else v1


def _sig_base_row(sig_idx: int) -> int:
    """신호 인덱스를 파형 시작 행 번호(H행)로 변환"""
    return ROW_WAVE_START + sig_idx * (CELLS_PER_SIGNAL + SIGNAL_GAP_ROWS)


def _sig_timing_row(sig_idx: int) -> int:
    """신호 인덱스를 timing 라벨 행 번호로 변환 (H행 바로 위)"""
    return _sig_base_row(sig_idx) - 1


# ────────────────────────────────────────────────────────────────
# 메인 익스포터
# ────────────────────────────────────────────────────────────────

class ExcelWaveformExporter:
    """
    Excel 파형 시각화 내보내기 클래스

    openpyxl을 사용하여 신호 파형을 셀 테두리로 시각화합니다.
    """

    def export_all_models(self, filepath: str, model_store) -> bool:
        """
        model_store의 전체 모델을 각 시트에 파형으로 내보내기
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl이 필요합니다. 'pip install openpyxl'을 실행하세요.")

        if not model_store.models:
            return False

        wb = Workbook()
        first_sheet = True

        for md in model_store.models:
            sig_dicts = []
            for i, sig in enumerate(md.signals):
                if hasattr(sig, 'to_dict'):
                    d = sig.to_dict()
                    d['num'] = getattr(sig, '_num', f'S{i+1:02d}')
                else:
                    d = dict(sig)
                sig_dicts.append(d)

            # 모든 값이 0인 신호(visible=False) 제외
            visible = [s for s in sig_dicts if s.get('visible', True)]
            if not visible:
                continue

            sync_data_us = md.sync_data_us
            model_name   = md.name or md.model_num

            if first_sheet:
                ws = wb.active
                safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
                ws.title = safe_title
                first_sheet = False
            else:
                safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
                ws = wb.create_sheet(title=safe_title)

            self._draw_sheet(ws, visible, sync_data_us, model_name)

        wb.save(filepath)
        return True

    def export(self, filepath: str, signals: List[Dict],
               sync_data_us: float, model_name: str = "Model") -> bool:
        """단일 모델 신호 파형을 Excel 파일로 내보내기 (하위 호환용)"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("openpyxl이 필요합니다. 'pip install openpyxl'을 실행하세요.")

        if not signals:
            return False

        wb = Workbook()
        ws = wb.active
        safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
        ws.title = safe_title
        ws.sheet_view.zoomScale = ZOOM_PERCENT

        self._draw_sheet(ws, signals, sync_data_us, model_name)

        wb.save(filepath)
        return True

    def _draw_sheet(self, ws, signals: List[Dict],
                    sync_data_us: float, model_name: str):
        """워크시트에 파형 그리기"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.sheet_view.zoomScale = ZOOM_PERCENT

        # ── 구간 계산 ─────────────────────────────────────────────
        segments = _compute_segments(sync_data_us, signals)
        n_segs   = len(segments)
        total_us = sync_data_us

        TOTAL_WAVE_COLS = min(n_segs * 10, 160)
        seg_cols = []
        for s, e, _ in segments:
            ratio = (e - s) / total_us if total_us > 0 else 1 / n_segs
            seg_cols.append(max(1, round(ratio * TOTAL_WAVE_COLS)))

        diff = TOTAL_WAVE_COLS - sum(seg_cols)
        seg_cols[-1] = max(1, seg_cols[-1] + diff)

        seg_start_cols = []
        cur = COL_WAVE_START
        for n in seg_cols:
            seg_start_cols.append(cur)
            cur += n
        total_wave_end_col = cur - 1

        # ── 열 너비 설정 ───────────────────────────────────────────
        ws.column_dimensions[get_column_letter(COL_NUM)].width  = NUM_COL_WIDTH
        ws.column_dimensions[get_column_letter(COL_NAME)].width = NAME_COL_WIDTH
        for c in range(COL_WAVE_START, total_wave_end_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = WAVE_COL_WIDTH

        # ── 행 높이 설정 ───────────────────────────────────────────
        ws.row_dimensions[ROW_LABEL].height = 22
        # 첫 신호의 timing행 (row 2) 높이
        ws.row_dimensions[ROW_WAVE_START - 1].height = 14

        for sig_idx in range(len(signals)):
            base = _sig_base_row(sig_idx)
            timing_row = _sig_timing_row(sig_idx)
            ws.row_dimensions[timing_row].height = 14
            for r in range(base, base + CELLS_PER_SIGNAL):
                ws.row_dimensions[r].height = 10
            # 여백 행 (gap의 첫 번째 행만 여백, 두 번째는 다음 신호 timing행)
            gap_row1 = base + CELLS_PER_SIGNAL
            ws.row_dimensions[gap_row1].height = 4

        # ── 스타일 헬퍼 ────────────────────────────────────────────
        center_align = Alignment(horizontal='center', vertical='center')

        def solid_side(style='thin', color='FF000000'):
            return Side(style=style, color=color)

        header_fill = PatternFill('solid', fgColor=LABEL_COLOR)
        label_font  = Font(bold=True, size=FONT_LABEL)

        # ── 행 1: 구간 라벨 ─────────────────────────────────────
        ws.cell(ROW_LABEL, COL_NUM,  "NUM").font  = label_font
        ws.cell(ROW_LABEL, COL_NAME, "신호 이름").font = label_font
        ws.cell(ROW_LABEL, COL_NUM).fill  = header_fill
        ws.cell(ROW_LABEL, COL_NAME).fill = header_fill
        ws.cell(ROW_LABEL, COL_NUM).alignment  = center_align
        ws.cell(ROW_LABEL, COL_NAME).alignment = center_align

        for si, (start, end, label) in enumerate(segments):
            c1 = seg_start_cols[si]
            c2 = c1 + seg_cols[si] - 1
            cell = ws.cell(ROW_LABEL, c1, label)
            cell.font      = label_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = Border(
                left=solid_side('medium'), right=solid_side('medium'),
                top=solid_side('medium'),  bottom=solid_side('medium'))
            if c2 > c1:
                ws.merge_cells(start_row=ROW_LABEL, start_column=c1,
                               end_row=ROW_LABEL, end_column=c2)

        # ── 신호별 파형 그리기 ────────────────────────────────────
        for sig_idx, sig in enumerate(signals):
            base  = _sig_base_row(sig_idx)
            h_row = base           # High 행
            m_row = base + 1       # Mid 행
            l_row = base + 2       # Low 행

            # NUM / NAME 셀
            num_str = sig.get('num', f'S{sig_idx + 1:02d}')
            ws.cell(h_row, COL_NUM, num_str).font = Font(bold=True, size=9)
            ws.cell(h_row, COL_NUM).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=COL_NUM,
                           end_row=l_row,   end_column=COL_NUM)

            ws.cell(h_row, COL_NAME, sig.get('name', '')).font = Font(bold=True, size=FONT_NAME)
            ws.cell(h_row, COL_NAME).alignment = center_align
            ws.merge_cells(start_row=h_row, start_column=COL_NAME,
                           end_row=l_row,   end_column=COL_NAME)

            # 구간별 파형 셀 그리기
            prev_voltage = None
            for si_idx, (start, end, _) in enumerate(segments):
                mid_t   = (start + end) / 2
                voltage = _get_level(sig, mid_t)
                c1 = seg_start_cols[si_idx]
                c2 = c1 + seg_cols[si_idx] - 1

                is_transition = (prev_voltage is not None and
                                 abs(prev_voltage - voltage) > 1e-6)

                self._draw_waveform_cells(
                    ws, h_row, m_row, l_row, c1, c2,
                    voltage, prev_voltage, is_transition,
                    solid_side
                )

                # 첫 구간에 전압값 레이블 표시 (V 단위)
                if si_idx == 0:
                    v_label = f"{voltage:+.1f}V"
                    if voltage == 0:
                        vc = ws.cell(m_row, c1, v_label)
                        vc.font = Font(color='FF00AA00', size=FONT_VOLTAGE, bold=True)
                    elif voltage > 0:
                        vc = ws.cell(h_row, c1, v_label)
                        vc.font = Font(color='FFCC0000', size=FONT_VOLTAGE, bold=True)
                    else:
                        vc = ws.cell(l_row, c1, v_label)
                        vc.font = Font(color='FF0033CC', size=FONT_VOLTAGE, bold=True)

                prev_voltage = voltage

            # timing 정보 표시
            self._draw_signal_timing(
                ws, sig, sig_idx, segments, seg_start_cols, seg_cols,
                sync_data_us, center_align, solid_side
            )

    def _draw_waveform_cells(self, ws, h_row, m_row, l_row,
                              c1, c2, voltage, prev_voltage,
                              is_transition, solid_side_fn):
        """
        파형 셀 테두리 그리기

        - 전환(transition) 구간의 c1 열에 H/M/L 모두 굵은 left border (수직 연결선)
        - 비전환 구간 경계에는 border 없음 (깔끔한 파형 표시)
        - 수평 파형선: 전압 레벨에 따라 H/M/L 중 해당 행에만 그림
        """
        from openpyxl.styles import Border, Side

        thick = solid_side_fn('thick')
        none  = Side(style=None)

        for col in range(c1, c2 + 1):
            is_left_col = (col == c1)

            # 좌측 경계선: 전환 시에만 굵은선, 비전환은 없음
            if is_transition and is_left_col:
                left_h = thick
                left_m = thick
                left_l = thick
            else:
                left_h = none
                left_m = none
                left_l = none

            # 수평 파형선
            if voltage > 0:
                ws.cell(h_row, col).border = Border(top=thick,    left=left_h)
                ws.cell(m_row, col).border = Border(left=left_m)
                ws.cell(l_row, col).border = Border(left=left_l)
            elif voltage == 0:
                ws.cell(h_row, col).border = Border(left=left_h)
                ws.cell(m_row, col).border = Border(top=thick,    left=left_m)
                ws.cell(l_row, col).border = Border(left=left_l)
            else:
                ws.cell(h_row, col).border = Border(left=left_h)
                ws.cell(m_row, col).border = Border(left=left_m)
                ws.cell(l_row, col).border = Border(bottom=thick, left=left_l)

    def _draw_signal_timing(self, ws, sig: Dict, sig_idx: int,
                             segments, seg_start_cols, seg_cols,
                             sync_data_us: float,
                             center_align, solid_side_fn):
        """
        각 신호의 timing 정보를 파형 윗행(timing_row)에 구분하여 표시.

        표시 항목:
          - Delay 구간 (0 → delay):        "D: Xus"  보라색, 해당 열 범위 중앙
          - Width 구간 (delay → delay+width): "W: Xus"  파란색, 해당 열 범위 중앙
          - Period (0 → period):            "T: Xus"  주황색, period 범위 우측
        """
        from openpyxl.styles import Font

        timing_row = _sig_timing_row(sig_idx)

        delay  = float(sig.get('delay',  0))
        width  = float(sig.get('width',  0))
        period = float(sig.get('period', 0))

        if delay == 0 and width == 0 and period == 0:
            return

        total_cols = sum(seg_cols)

        def us_to_col(us_val: float) -> int:
            """us 값을 Excel 열 번호로 변환"""
            ratio = us_val / sync_data_us if sync_data_us > 0 else 0
            return COL_WAVE_START + int(ratio * total_cols)

        def write_timing_label(us_start, us_end, text, color):
            """timing_row의 지정 범위 중앙에 라벨 작성"""
            c_start = us_to_col(us_start)
            c_end   = max(c_start, us_to_col(us_end) - 1)
            mid_col = (c_start + c_end) // 2
            # 유효 열 범위 클램프
            max_col = COL_WAVE_START + total_cols - 1
            mid_col = max(COL_WAVE_START, min(mid_col, max_col))
            cell = ws.cell(timing_row, mid_col)
            cell.value     = text
            cell.font      = Font(color=color, size=FONT_TIMING, bold=True)
            cell.alignment = center_align

        # Delay 구간 표시
        if delay > 0:
            write_timing_label(0, delay, f"D: {_format_us(delay)}", 'FF7030A0')

        # Width 구간 표시
        if width > 0:
            write_timing_label(delay, delay + width, f"W: {_format_us(width)}", 'FF0070C0')

        # Period 표시 (delay+width 이후의 나머지 범위 우측)
        if period > 0:
            # period 전체 범위 표시 (겹치지 않도록 period 우측에 배치)
            p_col_end = us_to_col(period) - 1
            max_col   = COL_WAVE_START + total_cols - 1
            p_col_end = max(COL_WAVE_START, min(p_col_end, max_col))
            cell = ws.cell(timing_row, p_col_end)
            if cell.value is None:
                cell.value     = f"T: {_format_us(period)}"
                cell.font      = Font(color='FFCC5500', size=FONT_TIMING, bold=True)
                cell.alignment = center_align

    def _us_range_to_cols(self, us_start: float, us_end: float,
                           sync_data_us: float,
                           segments, seg_start_cols, seg_cols
                           ) -> Optional[Tuple[int, int]]:
        """us 시간 범위를 Excel 열 범위로 변환"""
        if us_end <= us_start or sync_data_us <= 0:
            return None
        total_cols = sum(seg_cols)
        c_start = COL_WAVE_START + int(us_start / sync_data_us * total_cols)
        c_end   = COL_WAVE_START + int(us_end   / sync_data_us * total_cols) - 1
        if c_start > c_end:
            c_end = c_start
        return c_start, c_end
