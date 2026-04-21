"""
제어 패널 (Control Panel)

상단 고정 패널로 다음 기능을 제공합니다.
  - 뷰 제어: 프레임 수, X축 모드, 뷰 시간 설정
  - 기타 설정: 그리드 토글, 뷰 모드(개별/합쳐 보기), 범례 위치
  - OTD/Excel I/O: 불러오기(OTD/Excel), 내보내기(OTD/Excel 파형/Excel 데이터),
                   포맷 파일 생성

변경 이력:
  v3: 모델 설정 패널(디스플레이 모델/주파수) 제거
  v3: OTD/Excel 불러오기 model_store 기반으로 통일
  v4 (이번 수정):
    - 상단 3개 LabelFrame → 1줄 컴팩트 Toolbar로 통합 (공간 절약)
    - Excel 파형 출력 → model_store 전체 모델을 시트별로 저장
    - Excel 데이터 출력 → model_store 전체 모델을 시트별로 저장
    - 전압/시간 단위 V, us 통일
"""

import tkinter as tk
from tkinter import ttk, messagebox
import os


class ControlPanel(tk.Frame):
    """
    제어 패널 위젯 클래스

    애플리케이션 상단에 배치되어 뷰 제어 및 파일 I/O 기능을 제공합니다.
    1줄 컴팩트 Toolbar 형태로 공간 최소화.

    Attributes:
        sync_data_manager: SyncData 관리자
        timing_viewer: 타이밍 다이어그램 뷰어 (지연 연결)
        signal_manager: 신호 관리자
        signal_storage: 신호 저장소
        pattern_data_panel: 패턴 데이터 패널 (지연 연결)
        model_store: 다중 모델 데이터 저장소
    """

    def __init__(self, parent, sync_data_manager, timing_viewer, signal_manager,
                 signal_storage, pattern_data_panel=None, model_store=None):
        super().__init__(parent, bg='#f0f0f0')

        self.sync_data_manager  = sync_data_manager
        self.timing_viewer      = timing_viewer
        self.signal_manager     = signal_manager
        self.signal_storage     = signal_storage
        self.pattern_data_panel = pattern_data_panel
        self.model_store        = model_store

        self._setup_ui()

    def _setup_ui(self):
        """UI 구성 — 1줄 컴팩트 Toolbar"""

        # ── 공통 스타일 ───────────────────────────────────────────
        toolbar_bg   = '#f0f0f0'
        section_bg   = '#e8e8e8'
        label_kw     = dict(bg=section_bg, fg='#333333', font=('Arial', 8))
        sep_kw       = dict(bg='#cccccc', width=1)

        def _sep():
            """수직 구분선"""
            tk.Frame(self, **sep_kw).pack(side=tk.LEFT, fill=tk.Y, padx=4, pady=4)

        # ── 섹션 1: 뷰 제어 ──────────────────────────────────────
        sec1 = tk.Frame(self, bg=section_bg, relief=tk.GROOVE, bd=1)
        sec1.pack(side=tk.LEFT, fill=tk.Y, padx=(6, 0), pady=4)

        tk.Label(sec1, text="뷰 제어", bg=section_bg, fg='#555555',
                 font=('Arial', 7, 'bold')).pack(anchor='w', padx=4, pady=(2, 0))

        row1 = tk.Frame(sec1, bg=section_bg)
        row1.pack(side=tk.TOP, fill=tk.X, padx=4, pady=(0, 4))

        # 프레임 수
        tk.Label(row1, text="프레임:", **label_kw).pack(side=tk.LEFT)
        self.frame_spinbox = tk.Spinbox(row1, from_=1, to=10,
                                         font=('Arial', 8), width=3,
                                         command=self._on_frame_changed,
                                         bg='#ffffff')
        self.frame_spinbox.delete(0, tk.END)
        self.frame_spinbox.insert(0, '2')
        self.frame_spinbox.pack(side=tk.LEFT, padx=(2, 8))

        # X축 모드
        tk.Label(row1, text="X축:", **label_kw).pack(side=tk.LEFT)
        self.x_axis_mode_var = tk.StringVar(value="frame")
        ttk.Radiobutton(row1, text="Frame", variable=self.x_axis_mode_var,
                        value="frame", command=self._on_x_axis_mode_changed).pack(side=tk.LEFT)
        ttk.Radiobutton(row1, text="Time(us)", variable=self.x_axis_mode_var,
                        value="time", command=self._on_x_axis_mode_changed).pack(side=tk.LEFT)

        # 뷰 시간
        tk.Label(row1, text=" 시간(us):", **label_kw).pack(side=tk.LEFT)
        self.view_time_entry = tk.Entry(row1, font=('Arial', 8), width=6, bg='#ffffff')
        self.view_time_entry.pack(side=tk.LEFT, padx=(2, 0))
        self.view_time_entry.bind('<Return>',   self._on_view_time_changed)
        self.view_time_entry.bind('<FocusOut>', self._on_view_time_changed)

        _sep()

        # ── 섹션 2: 기타 설정 ────────────────────────────────────
        sec2 = tk.Frame(self, bg=section_bg, relief=tk.GROOVE, bd=1)
        sec2.pack(side=tk.LEFT, fill=tk.Y, padx=0, pady=4)

        tk.Label(sec2, text="기타 설정", bg=section_bg, fg='#555555',
                 font=('Arial', 7, 'bold')).pack(anchor='w', padx=4, pady=(2, 0))

        row2 = tk.Frame(sec2, bg=section_bg)
        row2.pack(side=tk.TOP, fill=tk.X, padx=4, pady=(0, 4))

        # 그리드 토글
        tk.Button(row2, text="그리드", command=self._on_toggle_grid,
                  bg='#d0d0d0', fg='#333333', font=('Arial', 8, 'bold'),
                  relief=tk.RAISED, bd=1, padx=5, pady=2).pack(side=tk.LEFT, padx=(0, 6))

        # 뷰 모드
        tk.Label(row2, text="보기:", **label_kw).pack(side=tk.LEFT)
        self.view_mode_var = tk.StringVar(value="separate")
        ttk.Radiobutton(row2, text="개별", variable=self.view_mode_var,
                        value="separate", command=self._on_view_mode_changed).pack(side=tk.LEFT)
        ttk.Radiobutton(row2, text="합쳐", variable=self.view_mode_var,
                        value="combined", command=self._on_view_mode_changed).pack(side=tk.LEFT)

        # 범례 위치
        tk.Label(row2, text=" 범례:", **label_kw).pack(side=tk.LEFT)
        self.legend_combo = ttk.Combobox(row2, font=('Arial', 8),
                                         state='readonly', width=10)
        self.legend_combo['values'] = [
            '우상(upper right)', '좌상(upper left)',
            '우하(lower right)', '좌하(lower left)'
        ]
        self.legend_combo.current(0)
        self.legend_combo.pack(side=tk.LEFT, padx=(2, 0))
        self.legend_combo.bind('<<ComboboxSelected>>', self._on_legend_location_changed)

        _sep()

        # ── 섹션 3: OTD / Excel I/O ──────────────────────────────
        sec3 = tk.Frame(self, bg=section_bg, relief=tk.GROOVE, bd=1)
        sec3.pack(side=tk.LEFT, fill=tk.Y, padx=0, pady=4)

        tk.Label(sec3, text="OTD / Excel", bg=section_bg, fg='#555555',
                 font=('Arial', 7, 'bold')).pack(anchor='w', padx=4, pady=(2, 0))

        row3 = tk.Frame(sec3, bg=section_bg)
        row3.pack(side=tk.TOP, fill=tk.X, padx=4, pady=(0, 4))

        btn_data = [
            ("OTD 불러오기",       self._on_load_otd,              '#5a8a5a', '#ffffff'),
            ("Excel 불러오기",     self._on_load_excel,            '#5a7a9a', '#ffffff'),
            ("OTD 내보내기",       self._on_export_otd,            '#8a7a5a', '#ffffff'),
            ("Excel 데이터 내보내기", self._on_export_excel,        '#5a8a8a', '#ffffff'),
            ("Excel 파형 내보내기",  self._on_export_excel_waveform, '#7a5a9a', '#ffffff'),
            ("포맷 생성",          self._on_create_format,         '#7a7a7a', '#ffffff'),
        ]
        for text, cmd, bg_color, fg_color in btn_data:
            tk.Button(row3, text=text, command=cmd,
                      bg=bg_color, fg=fg_color,
                      font=('Arial', 8, 'bold'),
                      relief=tk.RAISED, bd=1, padx=6, pady=3,
                      cursor='hand2').pack(side=tk.LEFT, padx=2)

    # ──────────────────────────────────────────────────────────────
    # 뷰 제어 핸들러
    # ──────────────────────────────────────────────────────────────

    def _on_frame_changed(self):
        try:
            num_frames = int(self.frame_spinbox.get())
            if self.timing_viewer:
                self.timing_viewer.set_num_frames(num_frames)
        except (ValueError, AttributeError):
            pass

    def _on_x_axis_mode_changed(self):
        if self.timing_viewer:
            self.timing_viewer.set_x_axis_mode(self.x_axis_mode_var.get())

    def _on_view_time_changed(self, event=None):
        try:
            val = self.view_time_entry.get().strip()
            view_time = float(val) if val else None
            if self.timing_viewer:
                self.timing_viewer.set_view_time(view_time)
        except (ValueError, AttributeError):
            pass

    def _on_legend_location_changed(self, event=None):
        location_map = {
            '우상(upper right)': 'upper right',
            '좌상(upper left)':  'upper left',
            '우하(lower right)': 'lower right',
            '좌하(lower left)':  'lower left',
        }
        location = location_map.get(self.legend_combo.get(), 'upper right')
        if self.timing_viewer:
            self.timing_viewer.set_legend_location(location)

    def _on_toggle_grid(self):
        if self.timing_viewer:
            self.timing_viewer.toggle_grid()

    def _on_view_mode_changed(self):
        if self.timing_viewer:
            self.timing_viewer.set_view_mode(self.view_mode_var.get())

    # ──────────────────────────────────────────────────────────────
    # OTD 불러오기
    # ──────────────────────────────────────────────────────────────

    def _on_load_otd(self):
        """OTD 파일 불러오기 — 모든 모델을 model_store에 저장"""
        from tkinter import filedialog
        try:
            from otd_parser import OtdParser
        except ImportError as e:
            messagebox.showerror("오류", f"OTD 파서 로드 실패:\n{e}")
            return

        filepath = filedialog.askopenfilename(
            filetypes=[("OTD files", "*.otd *.OTD"), ("All files", "*.*")],
            title="OTD 파일 불러오기"
        )
        if not filepath:
            return

        try:
            otd_file = OtdParser.parse(filepath)
        except Exception as e:
            messagebox.showerror("파싱 오류", f"OTD 파일 파싱 실패:\n{e}")
            return

        if not otd_file.models:
            messagebox.showwarning("경고", "OTD 파일에서 모델을 찾을 수 없습니다.")
            return

        try:
            from otd_parser import OtdParser, otd_file_to_model_store
            model_list, mrt_groups = otd_file_to_model_store(otd_file)
        except Exception as e:
            messagebox.showerror("변환 오류", f"모델 데이터 변환 실패:\n{e}")
            return

        self.model_store.set_models(model_list, mrt_groups)

        n_models = len(model_list)
        n_mrt    = len(mrt_groups)
        messagebox.showinfo(
            "OTD 불러오기 완료",
            f"파일: {os.path.basename(filepath)}\n"
            f"모델: {n_models}개  |  MULTIREMOTE: {n_mrt}개\n\n"
            f"좌측 모델 목록에서 모델을 클릭하세요."
        )

    # ──────────────────────────────────────────────────────────────
    # Excel 불러오기
    # ──────────────────────────────────────────────────────────────

    def _on_load_excel(self):
        """Excel PG Signal 파일 불러오기 — 결과를 model_store에 저장"""
        from tkinter import filedialog
        try:
            from excel_importer import import_excel_all_models
        except ImportError as e:
            messagebox.showerror("오류", f"Excel 임포터 로드 실패:\n{e}")
            return

        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            title="PG Signal Excel 파일 불러오기"
        )
        if not filepath:
            return

        try:
            model_list = import_excel_all_models(filepath)
        except Exception as e:
            messagebox.showerror("오류", f"Excel 파싱 실패:\n{e}")
            return

        if not model_list:
            messagebox.showwarning("경고", "불러온 모델이 없습니다.\nExcel 형식을 확인하세요.")
            return

        self.model_store.set_models(model_list, [])

        n_sigs_total = sum(len(m.signals) for m in model_list)
        messagebox.showinfo(
            "Excel 불러오기 완료",
            f"파일: {os.path.basename(filepath)}\n"
            f"모델: {len(model_list)}개  |  총 신호: {n_sigs_total}개\n\n"
            f"좌측 모델 목록에서 모델을 클릭하세요."
        )

    # ──────────────────────────────────────────────────────────────
    # OTD 내보내기
    # ──────────────────────────────────────────────────────────────

    def _on_export_otd(self):
        """현재 model_store의 모든 모델을 OTD 파일로 내보내기"""
        from tkinter import filedialog
        try:
            from otd_exporter import OtdExporter
        except ImportError as e:
            messagebox.showerror("오류", f"OTD 내보내기 모듈 로드 실패:\n{e}")
            return

        if self.model_store and self.model_store.models:
            filepath = filedialog.asksaveasfilename(
                defaultextension=".otd",
                filetypes=[("OTD files", "*.otd"), ("All files", "*.*")],
                initialfile="models_export.otd",
                title="OTD 파일 저장"
            )
            if not filepath:
                return

            exporter = OtdExporter()
            if exporter.export_from_model_store(filepath, self.model_store):
                messagebox.showinfo("완료", f"OTD 파일 저장:\n{filepath}")
            else:
                messagebox.showerror("오류", "OTD 파일 저장에 실패했습니다.")
            return

        # model_store가 비었으면 현재 signal_manager 신호만 내보내기
        signals = self.signal_manager.get_all_signals()
        if not signals:
            messagebox.showwarning("경고", "내보낼 신호가 없습니다.")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".otd",
            filetypes=[("OTD files", "*.otd"), ("All files", "*.*")],
            initialfile="export.otd",
            title="OTD 파일 저장"
        )
        if not filepath:
            return

        sync_val = self.sync_data_manager.get_current_sync_data() * 1_000_000
        patterns = []
        if self.pattern_data_panel is not None:
            patterns = self.pattern_data_panel.get_patterns()

        model_data = [{
            'model_num':    '001',
            'name':         self.sync_data_manager.current_model or 'MODEL',
            'frequency_hz': float(self.sync_data_manager.current_frequency or 60),
            'sync_data_us': sync_val,
            'sync_cntr':    0,
            'signals':      [s.to_dict() for s in signals],
            'patterns':     patterns,
        }]

        exporter = OtdExporter()
        if exporter.export(filepath, model_data):
            messagebox.showinfo("완료", f"OTD 파일 저장:\n{filepath}")
        else:
            messagebox.showerror("오류", "OTD 파일 저장에 실패했습니다.")

    # ──────────────────────────────────────────────────────────────
    # Excel 파형 출력 — 전체 모델을 시트별로 저장
    # ──────────────────────────────────────────────────────────────

    def _on_export_excel_waveform(self):
        """
        Excel 파형 시각화 내보내기

        model_store에 모델이 있으면 전체 모델을 각 시트에 저장.
        없으면 현재 signal_manager 신호만 저장.
        단위: 전압=V, 시간=us
        """
        from tkinter import filedialog
        try:
            from excel_waveform_exporter import ExcelWaveformExporter
        except ImportError as e:
            messagebox.showerror("오류", f"Excel 파형 내보내기 모듈 로드 실패:\n{e}")
            return

        exporter = ExcelWaveformExporter()

        # model_store에 여러 모델이 있으면 전체 저장
        if self.model_store and len(self.model_store.models) > 0:
            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="all_models_waveform.xlsx",
                title="Excel 파형 시각화 저장 (전체 모델)"
            )
            if not filepath:
                return
            try:
                exporter.export_all_models(filepath, self.model_store)
                messagebox.showinfo(
                    "완료",
                    f"Excel 파형 저장 완료:\n{filepath}\n"
                    f"모델 수: {len(self.model_store.models)}개 (각 시트별)"
                )
            except Exception as e:
                messagebox.showerror("오류", f"Excel 파형 저장 실패:\n{e}")
            return

        # model_store 없음 → 현재 신호만
        signals = self.signal_manager.get_all_signals()
        visible_signals = [s for s in signals if getattr(s, 'visible', True)]
        if not visible_signals:
            messagebox.showwarning("경고", "내보낼 신호가 없습니다.")
            return

        model        = self.sync_data_manager.current_model or 'MODEL'
        sync_data_s  = self.sync_data_manager.get_current_sync_data()
        sync_data_us = sync_data_s * 1_000_000

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{model}_waveform.xlsx",
            title="Excel 파형 시각화 저장"
        )
        if not filepath:
            return

        sig_dicts = [s.to_dict() for s in visible_signals]
        for i, (sig, d) in enumerate(zip(visible_signals, sig_dicts)):
            d['num'] = getattr(sig, '_num', f'S{i+1:02d}')

        try:
            exporter.export(filepath, sig_dicts, sync_data_us, model)
            messagebox.showinfo("완료", f"Excel 파형 시각화 저장:\n{filepath}")
        except Exception as e:
            messagebox.showerror("오류", f"Excel 파형 저장 실패:\n{e}")

    # ──────────────────────────────────────────────────────────────
    # 포맷 파일 생성
    # ──────────────────────────────────────────────────────────────

    def _on_create_format(self):
        """Excel 불러오기용 빈 양식 파일 생성"""
        from tkinter import filedialog, simpledialog
        try:
            from excel_importer import generate_excel_format_file
        except ImportError as e:
            messagebox.showerror("오류", f"양식 생성 모듈 로드 실패:\n{e}")
            return

        model_count = simpledialog.askinteger(
            "포맷 파일 생성",
            "생성할 빈 모델(시트) 수를 입력하세요:",
            initialvalue=1, minvalue=1, maxvalue=20
        )
        if model_count is None:
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="pg_signal_format.xlsx",
            title="Excel 포맷 파일 저장"
        )
        if not filepath:
            return

        try:
            generate_excel_format_file(filepath, model_count)
            messagebox.showinfo("완료", f"Excel 포맷 파일 생성:\n{filepath}")
        except Exception as e:
            messagebox.showerror("오류", f"포맷 파일 생성 실패:\n{e}")

    # ──────────────────────────────────────────────────────────────
    # Excel 데이터 출력 — 전체 모델을 시트별로 저장
    # ──────────────────────────────────────────────────────────────

    def _on_export_excel(self):
        """
        신호 데이터를 Excel 파일로 내보내기

        model_store에 모델이 있으면 전체 모델을 각 시트에 저장.
        단위: 전압=V, 시간=us

        출력 시트 구조:
          - [모델이름] 시트: 신호 파라미터 (excel_importer 양식 호환)
          - [Model Info] 시트: 모델 기본 정보 요약
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "오류",
                "openpyxl 라이브러리가 필요합니다.\n'pip install openpyxl'을 실행하세요."
            )
            return

        # model_store에 모델이 있으면 전체 저장
        if self.model_store and self.model_store.models:
            from tkinter import filedialog
            import re

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="all_models_data.xlsx",
                title="Excel 데이터 저장 (전체 모델)"
            )
            if not filepath:
                return

            try:
                self._export_all_models_to_excel(filepath)
                messagebox.showinfo(
                    "완료",
                    f"Excel 파일 저장:\n{filepath}\n"
                    f"모델 수: {len(self.model_store.models)}개 (각 시트별)"
                )
            except Exception as e:
                messagebox.showerror("오류", f"Excel 파일 저장 실패:\n{str(e)}")
            return

        # model_store 없음 → 현재 신호만
        self._export_current_signals_to_excel()

    def _export_all_models_to_excel(self, filepath: str):
        """model_store의 전체 모델을 Excel 시트별로 저장 (V, us 단위)"""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import re

        wb = Workbook()
        first = True

        for md in self.model_store.models:
            # Signal 객체 → dict 변환
            signals = []
            for sig in md.signals:
                if hasattr(sig, 'to_dict'):
                    signals.append(sig.to_dict())
                elif isinstance(sig, dict):
                    signals.append(sig)

            model_name   = md.name or md.model_num
            sync_data_us = md.sync_data_us
            freq_hz      = md.frequency_hz

            safe_title = re.sub(r'[\\/*?\[\]:]', '_', model_name)[:31]
            if first:
                ws = wb.active
                ws.title = safe_title
                first = False
            else:
                ws = wb.create_sheet(title=safe_title)

            self._write_signal_sheet(ws, signals, model_name,
                                     sync_data_us, freq_hz, md.patterns)

        # Model Info 시트
        ws_info = wb.create_sheet("Model Info")
        self._write_model_info_sheet(ws_info)

        wb.save(filepath)

    def _write_signal_sheet(self, ws, signals, model_name,
                             sync_data_us, freq_hz, patterns):
        """신호 파라미터 시트 작성 (V, us 단위)"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        header_fill  = PatternFill('solid', fgColor='FF2C3E50')
        sync_fill    = PatternFill('solid', fgColor='FF27AE60')
        row_fill_alt = PatternFill('solid', fgColor='FFECF0F1')

        header_font = Font(bold=True, size=10, color='FFFFFFFF')
        sync_font   = Font(bold=True, size=9,  color='FFFFFFFF')
        num_font    = Font(bold=True, size=9,  color='FF2C3E50')
        normal_font = Font(size=9)

        center = Alignment(horizontal='center', vertical='center')
        left   = Alignment(horizontal='left',   vertical='center')

        def side(style='thin'):
            return Side(style=style)

        thin_border = Border(left=side(), right=side(), top=side(), bottom=side())

        # 헤더 (단위: V, us)
        SIGNAL_HEADERS = [
            ('A', 'NUM',          8),
            ('B', 'NAME',         14),
            ('C', 'SIG TYPE',     9),
            ('D', 'SIG MODE',     9),
            ('E', 'INV',          7),
            ('F', 'V1 (V)',       9),
            ('G', 'V2 (V)',       9),
            ('H', 'V3 (V)',       9),
            ('I', 'V4 (V)',       9),
            ('J', 'DELAY (us)',   11),
            ('K', 'PERIOD (us)',  11),
            ('L', 'WIDTH (us)',   11),
            ('M', 'LENGTH (us)',  11),
            ('N', 'AREA (us)',    11),
        ]

        ws.row_dimensions[1].height = 20
        for col_letter, label, width in SIGNAL_HEADERS:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = thin_border
            ws.column_dimensions[col_letter].width = width

        for i in range(36):
            row_num = i + 2
            num_str = f'S{i + 1:02d}'
            num_cell = ws.cell(row=row_num, column=1, value=num_str)
            num_cell.font      = num_font
            num_cell.alignment = center
            num_cell.border    = thin_border
            if row_num % 2 == 0:
                num_cell.fill = row_fill_alt

            if i < len(signals):
                sig = signals[i]
                sig_values = [
                    sig.get('name', ''),           # B
                    sig.get('sig_type', ''),        # C
                    sig.get('sig_mode', 0),         # D
                    sig.get('inversion', 0),        # E
                    sig.get('v1', 0),               # F (V)
                    sig.get('v2', 0),               # G (V)
                    sig.get('v3', 0),               # H (V)
                    sig.get('v4', 0),               # I (V)
                    sig.get('delay', 0),            # J (us)
                    sig.get('period', 0),           # K (us)
                    sig.get('width', 0),            # L (us)
                    sig.get('length', sig.get('_length', 0)),  # M (us)
                    sig.get('area',   sig.get('_area',   0)),  # N (us)
                ]
            else:
                sig_values = [''] * 13

            for col_offset, val in enumerate(sig_values):
                col_idx = 2 + col_offset
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.border    = thin_border
                c.alignment = center
                c.font      = normal_font
                if row_num % 2 == 0:
                    c.fill = row_fill_alt

        # SyncData 영역 (P/Q열, 2~4행, us 단위)
        sync_labels = [
            ('SyncData (us)', sync_data_us),
            ('Frequency (Hz)', float(freq_hz)),
            ('SyncCounter', 0),
        ]
        ws.column_dimensions['P'].width = 16
        ws.column_dimensions['Q'].width = 12
        for row_offset, (label, val) in enumerate(sync_labels):
            row_num = row_offset + 2
            p_cell = ws.cell(row=row_num, column=16, value=label)
            p_cell.font      = sync_font
            p_cell.fill      = sync_fill
            p_cell.alignment = left
            p_cell.border    = thin_border

            q_cell = ws.cell(row=row_num, column=17, value=val)
            q_cell.border    = thin_border
            q_cell.alignment = center

        # 패턴 데이터 영역 (39~40행 이후)
        ws.row_dimensions[38].height = 6
        ws.row_dimensions[39].height = 18
        ws.row_dimensions[40].height = 18

        ws.merge_cells('A39:S39')
        sep = ws.cell(row=39, column=1, value='=== PATTERN DATA ===')
        sep.font      = Font(bold=True, size=9, color='FFFFFFFF')
        sep.fill      = PatternFill('solid', fgColor='FF8E44AD')
        sep.alignment = center

        PATTERN_HEADERS = [
            ('A', 'PTN_NO', 8), ('B', 'NAME',   12),
            ('C', 'R_V1',   8), ('D', 'R_V2',    8),
            ('E', 'R_V3',   8), ('F', 'R_V4',    8),
            ('G', 'G_V1',   8), ('H', 'G_V2',    8),
            ('I', 'G_V3',   8), ('J', 'G_V4',    8),
            ('K', 'B_V1',   8), ('L', 'B_V2',    8),
            ('M', 'B_V3',   8), ('N', 'B_V4',    8),
            ('O', 'W_V1',   8), ('P', 'W_V2',    8),
            ('Q', 'W_V3',   8), ('R', 'W_V4',    8),
            ('S', 'TYPE',   8),
        ]
        ptn_header_fill = PatternFill('solid', fgColor='FF8E44AD')
        ptn_header_font = Font(bold=True, size=9, color='FFFFFFFF')

        for col_letter, label, width in PATTERN_HEADERS:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=40, column=col_idx, value=label)
            cell.font      = ptn_header_font
            cell.fill      = ptn_header_fill
            cell.alignment = center
            cell.border    = thin_border

        for ptn_offset, ptn in enumerate(patterns or []):
            ptn_row = 41 + ptn_offset
            ptn_vals = [
                ptn.get('ptn_no', ptn_offset + 1),
                ptn.get('name', f'PTN{ptn_offset+1:02d}'),
                ptn.get('r_v1', 0), ptn.get('r_v2', 0),
                ptn.get('r_v3', 0), ptn.get('r_v4', 0),
                ptn.get('g_v1', 0), ptn.get('g_v2', 0),
                ptn.get('g_v3', 0), ptn.get('g_v4', 0),
                ptn.get('b_v1', 0), ptn.get('b_v2', 0),
                ptn.get('b_v3', 0), ptn.get('b_v4', 0),
                ptn.get('w_v1', 0), ptn.get('w_v2', 0),
                ptn.get('w_v3', 0), ptn.get('w_v4', 0),
                ptn.get('ptn_type', 0),
            ]
            for col_idx, val in enumerate(ptn_vals, 1):
                c = ws.cell(row=ptn_row, column=col_idx, value=val)
                c.border    = thin_border
                c.alignment = center

    def _write_model_info_sheet(self, ws_info):
        """Model Info 요약 시트 작성"""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        info_fill = PatternFill('solid', fgColor='FF2980B9')
        info_font = Font(bold=True, size=9, color='FFFFFFFF')
        center    = Alignment(horizontal='center', vertical='center')

        def side(style='thin'):
            return Side(style=style)
        thin_border = Border(left=side(), right=side(), top=side(), bottom=side())

        # 헤더
        for col_idx, label in enumerate(['Model', 'Name', 'Freq (Hz)', 'SyncData (us)'], 1):
            cell = ws_info.cell(row=1, column=col_idx, value=label)
            cell.font      = info_font
            cell.fill      = info_fill
            cell.alignment = center
            cell.border    = thin_border

        for row_offset, md in enumerate(self.model_store.models, 2):
            vals = [
                md.model_num,
                md.name,
                round(md.frequency_hz, 2),
                md.sync_data_us,
            ]
            for col_idx, val in enumerate(vals, 1):
                c = ws_info.cell(row=row_offset, column=col_idx, value=val)
                c.border    = thin_border
                c.alignment = center

        for col in ['A', 'B', 'C', 'D']:
            ws_info.column_dimensions[col].width = 16

    def _export_current_signals_to_excel(self):
        """현재 signal_manager 신호만 Excel로 내보내기 (model_store 없을 때 fallback)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("오류", "openpyxl이 필요합니다.")
            return

        from tkinter import filedialog
        import re

        signals = self.signal_manager.get_all_signals()
        visible_signals = [s for s in signals if getattr(s, 'visible', True)]
        if not visible_signals:
            messagebox.showwarning("경고", "내보낼 신호가 없습니다.")
            return

        model    = self.sync_data_manager.current_model or 'MODEL'
        freq     = self.sync_data_manager.current_frequency or 60
        sync_data_s  = self.sync_data_manager.get_current_sync_data()
        sync_data_us = sync_data_s * 1_000_000

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"{model}_data.xlsx",
            title="Excel 데이터 저장"
        )
        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            safe_title = re.sub(r'[\\/*?\[\]:]', '_', model)[:31]
            ws.title = safe_title

            sig_dicts = [s.to_dict() for s in visible_signals]
            patterns = []
            if self.pattern_data_panel is not None:
                patterns = self.pattern_data_panel.get_patterns()

            self._write_signal_sheet(ws, sig_dicts, model, sync_data_us, freq, patterns)
            wb.save(filepath)
            messagebox.showinfo("완료", f"Excel 파일 저장:\n{filepath}")
        except Exception as e:
            messagebox.showerror("오류", f"Excel 파일 저장 실패:\n{str(e)}")
