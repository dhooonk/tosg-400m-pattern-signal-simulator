"""
Excel 임포트 양식 파일 생성기 (피드백 13)

Excel 불러오기에 사용할 정해진 양식의 Excel 파일을 생성합니다.

양식 규격:
  A1=NUM, B1=NAME, C1=SIG TYPE, D1=SIG MODE, E1=INV
  F1=V1, G1=V2, H1=V3, I1=V4 (단위: V)
  J1=DELAY, K1=PERIOD, L1=WIDTH, M1=LENGTH, N1=AREA (단위: us)
  A2~A37: S01~S36
  P2=SyncData, Q2=(값), P3=Frequency, Q3=(Hz), P4=SyncCounter, Q4=(값)
  
  A40 이하: 패턴 데이터 섹션
  A39=== PATTERN DATA ===
  A40=PTN_NO, B40=NAME, C40=R_V1 ... R40=W_V4, S40=TYPE
"""

from typing import List


def generate_excel_format_file(filepath: str, model_count: int = 1) -> bool:
    """
    Excel 임포트용 빈 양식 파일 생성
    
    Args:
        filepath: 저장할 .xlsx 파일 경로
        model_count: 생성할 시트(모델) 수
        
    Returns:
        bool: 성공 여부
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl이 필요합니다.")

    wb = Workbook()
    # 기본 시트 제거 후 새로 생성
    wb.remove(wb.active)

    # ────────── 스타일 정의 ──────────────────────────────────
    header_fill  = PatternFill('solid', fgColor='FF2C3E50')  # 진한 남색
    sync_fill    = PatternFill('solid', fgColor='FF27AE60')  # 초록
    pattern_fill = PatternFill('solid', fgColor='FF8E44AD')  # 보라
    row_fill     = PatternFill('solid', fgColor='FFECF0F1')  # 연회색

    header_font  = Font(bold=True, size=10, color='FFFFFFFF')
    sync_font    = Font(bold=True, size=9, color='FFFFFFFF')
    pattern_font = Font(bold=True, size=9, color='FFFFFFFF')
    normal_font  = Font(size=9)
    num_font     = Font(bold=True, size=9, color='FF2C3E50')

    center = Alignment(horizontal='center', vertical='center')
    left   = Alignment(horizontal='left',   vertical='center')

    def _side(style='thin'):
        return Side(style=style)

    thin_border = Border(
        left=_side(), right=_side(), top=_side(), bottom=_side()
    )

    # ────────── 컬럼 헤더 정의 ──────────────────────────────
    SIGNAL_HEADERS = [
        ('A', 'NUM',      8),
        ('B', 'NAME',    14),
        ('C', 'SIG TYPE', 9),
        ('D', 'SIG MODE', 9),
        ('E', 'INV',      7),
        ('F', 'V1 (V)',   9), ('G', 'V2 (V)',   9),
        ('H', 'V3 (V)',   9), ('I', 'V4 (V)',   9),
        ('J', 'DELAY (us)',  11), ('K', 'PERIOD (us)', 11),
        ('L', 'WIDTH (us)',  11), ('M', 'LENGTH (us)', 11),
        ('N', 'AREA (us)',   11),
    ]

    PATTERN_HEADERS = [
        ('A', 'PTN_NO', 8), ('B', 'NAME',   12),
        ('C', 'R_V1',   8), ('D', 'R_V2',    8),
        ('E', 'R_V3',   8), ('F', 'R_V4',    8),
        ('G', 'G_V1',   8), ('H', 'G_V2',    8),
        ('I', 'G_V3',   8), ('J', 'G_V4',    8),
        ('K', 'B_V1',   8), ('L', 'B_V2',    8),
        ('M', 'B_V3',   8), ('N', 'B_V4',    8),
        ('O', 'W_V1',   8), ('P', 'W_V2',    8),
        ('Q', 'W_V3',   8), ('R', 'W_V4',    8),
        ('S', 'TYPE',   8),
    ]

    # ────────── 시트별 양식 생성 ─────────────────────────────
    for sheet_idx in range(1, model_count + 1):
        sheet_name = f"Model-{sheet_idx:03d}"
        ws = wb.create_sheet(title=sheet_name)

        # 행 높이
        ws.row_dimensions[1].height  = 20
        ws.row_dimensions[38].height = 6   # 빈 여백
        ws.row_dimensions[39].height = 18
        ws.row_dimensions[40].height = 18

        # ── 신호 헤더 (1행) ──────────────────────────────────
        for col_letter, label, width in SIGNAL_HEADERS:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = thin_border
            ws.column_dimensions[col_letter].width = width

        # ── 신호 행 S01~S36 (2~37행) ─────────────────────────
        for i in range(36):
            row = i + 2
            num_cell = ws.cell(row=row, column=1, value=f'S{i+1:02d}')
            num_cell.font      = num_font
            num_cell.alignment = center
            num_cell.border    = thin_border
            if row % 2 == 0:
                num_cell.fill = row_fill

            for col_idx in range(2, 15):  # B~N
                c = ws.cell(row=row, column=col_idx, value='')
                c.border    = thin_border
                c.alignment = center
                if row % 2 == 0:
                    c.fill = row_fill

        # ── SyncData 영역 (P/Q열, 2~4행) ─────────────────────
        sync_labels = ['SyncData (us)', 'Frequency (Hz)', 'SyncCounter']
        for i, label in enumerate(sync_labels):
            row = i + 2
            p_cell = ws.cell(row=row, column=16, value=label)  # P열
            p_cell.font      = sync_font
            p_cell.fill      = sync_fill
            p_cell.alignment = left
            p_cell.border    = thin_border
            ws.column_dimensions['P'].width = 16

            q_cell = ws.cell(row=row, column=17, value=0)  # Q열
            q_cell.border    = thin_border
            q_cell.alignment = center
            ws.column_dimensions['Q'].width = 12

        # ── 패턴 섹션 구분 (39행) ───────────────────────────
        ws.merge_cells('A39:S39')
        sep = ws.cell(row=39, column=1, value='=== PATTERN DATA (아래에 패턴 입력) ===')
        sep.font      = pattern_font
        sep.fill      = pattern_fill
        sep.alignment = center

        # ── 패턴 헤더 (40행) ────────────────────────────────
        for col_letter, label, width in PATTERN_HEADERS:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=40, column=col_idx, value=label)
            cell.font      = pattern_font
            cell.fill      = pattern_fill
            cell.alignment = center
            cell.border    = thin_border

        # ── 빈 패턴 행 (41~60행) ───────────────────────────
        for ptn_row in range(41, 61):
            for col_idx in range(1, 20):
                c = ws.cell(row=ptn_row, column=col_idx, value='')
                c.border    = thin_border
                c.alignment = center
                if ptn_row % 2 == 0:
                    c.fill = row_fill
            # A열: PTN No 기본값
            ptn_num = ptn_row - 40
            ws.cell(row=ptn_row, column=1, value=ptn_num).font = num_font

    wb.save(filepath)
    return True
